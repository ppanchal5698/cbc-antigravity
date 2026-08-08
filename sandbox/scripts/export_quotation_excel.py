import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

output_dir = os.environ.get("CBC_OUTPUTS_ROOT", "sandbox/outputs")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "CBC_Material_Quotation_Baldwin_PA.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "CBC Material Quotation"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

# Styling Palette - Modern Corporate Steel Blue & Navy
navy_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
subtotal_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
accent_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
total_fill = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")

white_title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
white_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
bold_font = Font(name="Calibri", size=11, bold=True, color="000000")
regular_font = Font(name="Calibri", size=11, bold=False, color="000000")
italic_font = Font(name="Calibri", size=10, italic=True, color="4A5568")

thin_border_side = Side(border_style="thin", color="CBD5E0")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
thick_bottom = Side(border_style="medium", color="1A365D")
double_bottom = Side(border_style="double", color="1A365D")
top_thin = Side(border_style="thin", color="1A365D")
subtotal_border = Border(top=top_thin, bottom=thick_bottom, left=thin_border_side, right=thin_border_side)
total_border = Border(top=top_thin, bottom=double_bottom, left=thin_border_side, right=thin_border_side)

# Header Title Block
ws.merge_cells("A1:I1")
ws["A1"] = "CONSTRUCTION BUILDING COMPONENTS (CBC)"
ws["A1"].font = white_title_font
ws["A1"].fill = navy_fill
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:I2")
ws["A2"] = "A Division of The Hamilton Parker Company | Commercial Material Quotation (DRAFT)"
ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
ws["A2"].fill = navy_fill
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 20

# Project Info Block
info = [
    ("Project:", "Baldwin PA Revision 4", "Quote Date:", "August 7, 2026"),
    ("Plan Set:", "Full Set 25-073 Baldwin PA Revision 4 set 4-14-26.pdf", "Project State:", "PA (0.0% Sales Tax - Supply-Only)"),
    ("Client Account:", "Standard Commercial General Bid", "Status:", "DRAFT — For Estimator Review"),
    ("Address:", "800 John C. Watts Dr., Nicholasville, KY 40356", "Phone / Fax:", "855-432-4613 | 877-887-7806")
]

row_idx = 4
for label1, val1, label2, val2 in info:
    ws.cell(row=row_idx, column=1, value=label1).font = bold_font
    ws.cell(row=row_idx, column=2, value=val1).font = regular_font
    ws.cell(row=row_idx, column=6, value=label2).font = bold_font
    ws.cell(row=row_idx, column=7, value=val2).font = regular_font
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

row_idx += 1 # Empty spacing row

# Table Headers
headers = ["Tag / Item", "Room / Section", "Description / Details", "Qty", "Unit", "Unit Sale ($)", "Ext Sale ($)", "Cost Sourcing Basis", "Plan & Catalog Citations"]
ws.row_dimensions[row_idx].height = 25
for col_idx, text in enumerate(headers, start=1):
    cell = ws.cell(row=row_idx, column=col_idx, value=text)
    cell.font = white_header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center" if col_idx in [4, 5] else ("right" if col_idx in [6, 7] else "left"), vertical="center")
    cell.border = thin_border

row_idx += 1

# Helper function to add a section header row
def add_section_header(title):
    global row_idx
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = bold_font
    cell.fill = accent_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = 22
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).border = thin_border
    row_idx += 1

# Section 1: Division 08 Doors & Hardware
add_section_header("DIVISION 08 — DOORS, FRAMES & ARCHITECTURAL HARDWARE")

door_rows = [
    ("01", "DINING", "3070 Aluminum Storefront Door & Frame Package", 1, "EA", 0.00, 0.00, "vendor_rfq (GC/LockNet pkg)", "[schedule] Sheet A1.1"),
    ("02", "DINING", "6070 Aluminum Storefront Pair Doors & Frame Package", 1, "EA", 0.00, 0.00, "vendor_rfq (GC/LockNet pkg)", "[schedule] Sheet A1.1"),
    ("03", "WASHING", "3670 16-Ga HM Door, HM Frame 5-7/8\", Cont. SS Hinge, Panic, Closer, Kick, Thresh, Sweep", 1, "EA", 672.16, 672.16, "catalog_list_x_multiplier", "[schedule] A1.1 / Hager #18"),
    ("04", "MEN", "3070 Wood Door, HM Frame 5-7/8\", Hinges (1.5 pr), Privacy Lock, Stop, Kick, Sweep, Hook", 1, "EA", 416.67, 416.67, "catalog_list_x_multiplier", "[schedule] A1.1 / Hager / Bobrick"),
    ("05", "WOMEN", "3070 Wood Door, HM Frame 5-7/8\", Hinges, Privacy Lock, Push/Pull, Stop, Kick, Sweep, Hook", 1, "EA", 461.58, 461.58, "catalog_list_x_multiplier", "[schedule] A1.1 / Trimco / Bobrick"),
    ("06", "OFFICE", "3070 Wood Door, HM Frame 5-7/8\", Hinges (1.5 pr), Yale Lockset, Stop, Sweep", 1, "EA", 404.97, 404.97, "manual_wholesaler_net", "[schedule] A1.1 / Wholesaler Net")
]

door_start_row = row_idx
for tag, room, desc, qty, unit, unit_sale, ext_sale, basis, citation in door_rows:
    ws.cell(row=row_idx, column=1, value=tag).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=2, value=room).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=3, value=desc).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=4, value=qty).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=5, value=unit).alignment = Alignment(horizontal="center", vertical="center")
    
    c6 = ws.cell(row=row_idx, column=6, value=unit_sale)
    c6.number_format = "$#,##0.00"
    c6.alignment = Alignment(horizontal="right", vertical="center")
    
    c7 = ws.cell(row=row_idx, column=7, value=f"=D{row_idx}*F{row_idx}")
    c7.number_format = "$#,##0.00"
    c7.alignment = Alignment(horizontal="right", vertical="center")
    
    ws.cell(row=row_idx, column=8, value=basis).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=9, value=citation).alignment = Alignment(horizontal="left", vertical="center")
    
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).font = regular_font
        ws.cell(row=row_idx, column=c).border = thin_border
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

door_end_row = row_idx - 1

# Door Subtotal Row
ws.cell(row=row_idx, column=3, value="Subtotal Division 08 Doors & Hardware:").font = bold_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
sub_cell = ws.cell(row=row_idx, column=7, value=f"=SUM(G{door_start_row}:G{door_end_row})")
sub_cell.font = bold_font
sub_cell.number_format = "$#,##0.00"
sub_cell.alignment = Alignment(horizontal="right", vertical="center")
for c in range(1, 10):
    ws.cell(row=row_idx, column=c).fill = subtotal_fill
    ws.cell(row=row_idx, column=c).border = subtotal_border
ws.row_dimensions[row_idx].height = 22
row_idx += 2

# Section 2: Division 10 Washroom Accessories
add_section_header("DIVISION 10 — WASHROOM ACCESSORIES")

acc_rows = [
    ("B-262", "PREP / RESTROOM", "Bobrick B-26212 Surface Mounted Stainless Steel Paper Towel Dispenser", 2, "EA", 81.14, 162.28, "catalog_list_x_multiplier", "[schedule] A2.1 / Bobrick p.8"),
    ("B-241", "PREP / OFFICE", "Soap Dispenser Wall Mount (Kay 3741 / Bobrick B-2111 equiv)", 2, "EA", 55.68, 111.36, "catalog_list_x_multiplier", "[schedule] A2.1 / Bobrick"),
    ("B-251", "PREP / OFFICE", "Sanitizer Dispenser Wall Mount (Kay 3741 / Bobrick equiv)", 2, "EA", 55.68, 111.36, "catalog_list_x_multiplier", "[schedule] A2.1 / Bobrick")
]

acc_start_row = row_idx
for tag, room, desc, qty, unit, unit_sale, ext_sale, basis, citation in acc_rows:
    ws.cell(row=row_idx, column=1, value=tag).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=2, value=room).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=3, value=desc).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=4, value=qty).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=5, value=unit).alignment = Alignment(horizontal="center", vertical="center")
    
    c6 = ws.cell(row=row_idx, column=6, value=unit_sale)
    c6.number_format = "$#,##0.00"
    c6.alignment = Alignment(horizontal="right", vertical="center")
    
    c7 = ws.cell(row=row_idx, column=7, value=f"=D{row_idx}*F{row_idx}")
    c7.number_format = "$#,##0.00"
    c7.alignment = Alignment(horizontal="right", vertical="center")
    
    ws.cell(row=row_idx, column=8, value=basis).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=9, value=citation).alignment = Alignment(horizontal="left", vertical="center")
    
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).font = regular_font
        ws.cell(row=row_idx, column=c).border = thin_border
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

acc_end_row = row_idx - 1

# Accessories Subtotal Row
ws.cell(row=row_idx, column=3, value="Subtotal Division 10 Washroom Accessories:").font = bold_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
sub_cell = ws.cell(row=row_idx, column=7, value=f"=SUM(G{acc_start_row}:G{acc_end_row})")
sub_cell.font = bold_font
sub_cell.number_format = "$#,##0.00"
sub_cell.alignment = Alignment(horizontal="right", vertical="center")
for c in range(1, 10):
    ws.cell(row=row_idx, column=c).fill = subtotal_fill
    ws.cell(row=row_idx, column=c).border = subtotal_border
ws.row_dimensions[row_idx].height = 22
row_idx += 2

# Section 3: Division 06 FRP Wall Panels & Trims
add_section_header("DIVISION 06 64 — FRP WALL PANELS & MOULDINGS (PROVISIONAL)")

frp_rows = [
    ("FRP-1", "KITCHEN/PREP", ".090\" Class C FRP Wall Panel 4'x10' White Embossed", 50, "SHT", 57.53, 2876.50, "catalog_list_x_multiplier", "[spec] Sheet A7.2 / NUDO"),
    ("TRIM-1", "KITCHEN/PREP", "10ft PVC Division Bar Moulding", 49, "STK", 8.90, 436.10, "catalog_list_x_multiplier", "[spec] Sheet A7.2 / NUDO"),
    ("TRIM-2", "KITCHEN/PREP", "10ft PVC Inside Corner Moulding", 12, "STK", 8.90, 106.80, "catalog_list_x_multiplier", "[spec] Sheet A7.2 / NUDO"),
    ("TRIM-3", "KITCHEN/PREP", "10ft PVC Outside Corner Moulding", 6, "STK", 8.90, 53.40, "catalog_list_x_multiplier", "[spec] Sheet A7.2 / NUDO"),
    ("TRIM-4", "KITCHEN/PREP", "10ft PVC End Cap J-Moulding", 36, "STK", 7.53, 271.08, "catalog_list_x_multiplier", "[spec] Sheet A7.2 / NUDO"),
    ("ADH-1", "KITCHEN/PREP", "FRP Water-Based Adhesive 4-Gallon Pails", 7, "PAIL", 93.15, 652.05, "catalog_list_x_multiplier", "[spec] Sheet A7.2 / NUDO")
]

frp_start_row = row_idx
for tag, room, desc, qty, unit, unit_sale, ext_sale, basis, citation in frp_rows:
    ws.cell(row=row_idx, column=1, value=tag).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=2, value=room).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=3, value=desc).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=4, value=qty).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=5, value=unit).alignment = Alignment(horizontal="center", vertical="center")
    
    c6 = ws.cell(row=row_idx, column=6, value=unit_sale)
    c6.number_format = "$#,##0.00"
    c6.alignment = Alignment(horizontal="right", vertical="center")
    
    c7 = ws.cell(row=row_idx, column=7, value=f"=D{row_idx}*F{row_idx}")
    c7.number_format = "$#,##0.00"
    c7.alignment = Alignment(horizontal="right", vertical="center")
    
    ws.cell(row=row_idx, column=8, value=basis).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=9, value=citation).alignment = Alignment(horizontal="left", vertical="center")
    
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).font = regular_font
        ws.cell(row=row_idx, column=c).border = thin_border
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

frp_end_row = row_idx - 1

# FRP Subtotal Row
ws.cell(row=row_idx, column=3, value="Subtotal Division 06 FRP Panels & Trims:").font = bold_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
sub_cell = ws.cell(row=row_idx, column=7, value=f"=SUM(G{frp_start_row}:G{frp_end_row})")
sub_cell.font = bold_font
sub_cell.number_format = "$#,##0.00"
sub_cell.alignment = Alignment(horizontal="right", vertical="center")
for c in range(1, 10):
    ws.cell(row=row_idx, column=c).fill = subtotal_fill
    ws.cell(row=row_idx, column=c).border = subtotal_border
ws.row_dimensions[row_idx].height = 22
row_idx += 2

# Summary Block
ws.cell(row=row_idx, column=6, value="Base Bid Material Subtotal:").font = bold_font
ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
base_cell = ws.cell(row=row_idx, column=7, value=f"=G{door_end_row+1}+G{acc_end_row+1}+G{frp_end_row+1}")
base_cell.font = bold_font
base_cell.number_format = "$#,##0.00"
base_cell.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[row_idx].height = 22
row_idx += 1

ws.cell(row=row_idx, column=6, value="Freight (F.O.B. Factory / CBC):").font = bold_font
ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
fr_cell = ws.cell(row=row_idx, column=7, value="TBD (Excluded)")
fr_cell.font = italic_font
fr_cell.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[row_idx].height = 20
row_idx += 1

ws.cell(row=row_idx, column=6, value="Pennsylvania Sales Tax (0.0%):").font = bold_font
ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
tax_cell = ws.cell(row=row_idx, column=7, value=0.00)
tax_cell.font = regular_font
tax_cell.number_format = "$#,##0.00"
tax_cell.alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[row_idx].height = 20
row_idx += 1

ws.cell(row=row_idx, column=6, value="GRAND TOTAL QUOTATION:").font = Font(name="Calibri", size=12, bold=True, color="1A365D")
ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
gt_cell = ws.cell(row=row_idx, column=7, value=f"=G{row_idx-3}+G{row_idx-1}")
gt_cell.font = Font(name="Calibri", size=12, bold=True, color="1A365D")
gt_cell.number_format = "$#,##0.00"
gt_cell.alignment = Alignment(horizontal="right", vertical="center")

for c in range(6, 8):
    ws.cell(row=row_idx, column=c).fill = total_fill
    ws.cell(row=row_idx, column=c).border = total_border
ws.row_dimensions[row_idx].height = 26
row_idx += 3

# Standard Terms & Conditions
add_section_header("STANDARD COMMERCIAL TERMS & CONDITIONS")
terms = [
    "1. Supply-Only material proposal (Material only, F.O.B. factory / CBC warehouse). Installation by others.",
    "2. Quotation is subject to CBC / The Hamilton Parker Company credit approval. Hamilton Parker PO required.",
    "3. Material prices are firm for 30 days from date of quotation.",
    "4. Freight: Excluded at estimate stage / carried TBD (billed at actual cost upon order release).",
    "5. Sales Tax: Pennsylvania 0% (Resale certificate / supply-only sale to General Contractor)."
]
for t in terms:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    ws.cell(row=row_idx, column=1, value=t).font = italic_font
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1

row_idx += 1
add_section_header("ESTIMATING NOTES & RFI REGISTER")
rfis = [
    "• Doors 01 & 02: Aluminum storefront doors & framing supplied by GC / LockNet package (carried as $0.00 direct material).",
    "• Division 06 FRP: Quantities carry 'provisional: true' flag (Open Item 5) — estimator must confirm conversion constants before ordering.",
    "• Door 06 Yale Lockset: Quoted at estimated wholesaler net ($85.00); confirm exact distributor quote from Banner/SecLock."
]
for r in rfis:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    ws.cell(row=row_idx, column=1, value=r).font = italic_font
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1

# Specific Column Width Adjustments
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 8
ws.column_dimensions["E"].width = 8
ws.column_dimensions["F"].width = 16
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 28
ws.column_dimensions["I"].width = 32

wb.save(output_path)
print(f"Quotation excel workbook saved successfully to {output_path}")
