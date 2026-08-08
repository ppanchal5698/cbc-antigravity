import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

output_dir = os.environ.get("CBC_OUTPUTS_ROOT", "sandbox/outputs")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "CBC_Material_Quotation_Architectural_Set.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "CBC Material Quotation"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

# Styling Palette - Modern Corporate Steel Blue & Navy
navy_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
subtotal_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
accent_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
total_fill = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")

white_title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
white_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
bold_font = Font(name="Calibri", size=11, bold=True, color="000000")
regular_font = Font(name="Calibri", size=11, bold=False, color="000000")
italic_font = Font(name="Calibri", size=10, italic=True, color="4A5568")

thin_border_side = Side(border_style="thin", color="CBD5E0")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
thick_bottom = Side(border_style="medium", color="1A365D")
double_bottom = Side(border_style="double", color="1A365D")
top_thin = Side(border_style="thin", color="1A365D")
subtotal_border = Border(top=top_thin, bottom=thick_bottom, left=thin_border_side, right=thin_border_side)
total_border = Border(top=top_thin, bottom=double_bottom, left=thin_border_side, right=thin_border_side)

# Header Title Block
ws.merge_cells("A1:I1")
ws["A1"] = "CONSTRUCTION BUILDING COMPONENTS (CBC)"
ws["A1"].font = white_title_font
ws["A1"].fill = navy_fill
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:I2")
ws["A2"] = "A Division of The Hamilton Parker Company | Commercial Material Quotation (DRAFT)"
ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
ws["A2"].fill = navy_fill
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 20

# Project Info Block
info = [
    ("Project:", "Commercial Building — 1_Architectural Set", "Quote Date:", "August 7, 2026"),
    ("Plan Set:", "1_Architectural.pdf (doc_id: 081cba06a10ab42a)", "Project State:", "PA (0.0% Sales Tax - Supply-Only)"),
    ("Client Account:", "Dutch Bros / Standard Commercial Account", "Status:", "DRAFT — For Estimator Review"),
    ("Address:", "800 John C. Watts Dr., Nicholasville, KY 40356", "Phone / Fax:", "855-432-4613 | 877-887-7806")
]

row_idx = 4
for label1, val1, label2, val2 in info:
    ws.cell(row=row_idx, column=1, value=label1).font = bold_font
    ws.cell(row=row_idx, column=2, value=val1).font = regular_font
    ws.cell(row=row_idx, column=6, value=label2).font = bold_font
    ws.cell(row=row_idx, column=7, value=val2).font = regular_font
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

row_idx += 1 # Empty spacing row

# Table Headers
headers = ["Tag / Item", "Room / Section", "Description / Details", "Qty", "Unit", "Unit Sale ($)", "Ext Sale ($)", "Cost Sourcing Basis", "Plan & Catalog Citations"]
ws.row_dimensions[row_idx].height = 25
for col_idx, text in enumerate(headers, start=1):
    cell = ws.cell(row=row_idx, column=col_idx, value=text)
    cell.font = white_header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center" if col_idx in [4, 5] else ("right" if col_idx in [6, 7] else "left"), vertical="center")
    cell.border = thin_border

row_idx += 1

# Helper function to add a section header row
def add_section_header(title):
    global row_idx
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = bold_font
    cell.fill = accent_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = 22
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).border = thin_border
    row_idx += 1

# Section 1: Division 08 Doors & Hardware
add_section_header("DIVISION 08 — DOORS, FRAMES & ARCHITECTURAL HARDWARE")

door_rows = [
    ("01", "BACK DOOR", "3670 16-Ga HM Door, 5-7/8\" HMD Frame, Ives Cont. Hinge, LCN 4040XP, Alarm Lock ETDL, Von Duprin 99EO, Ives Kick/Stop, Pemko 275A, Zero Sweep/Seal", 1, "EA", 2495.89, 2495.89, "catalog_list_x_multiplier", "[schedule] Sheet A2.2 / Ives, LCN, Alarm Lock, Von Duprin, Pemko, Zero"),
    ("02", "RUNNER DOOR", "3070 16-Ga HM Door, 5-7/8\" HMD Frame, Ives Cont. Hinge, LCN 4040XP, Alarm Lock ETDL, Von Duprin 99EO, Ives Kick/Stop, Pemko 275A, Zero Sweep/Seal", 1, "EA", 2390.40, 2390.40, "catalog_list_x_multiplier", "[schedule] Sheet A2.2 / Ives, LCN, Alarm Lock, Von Duprin, Pemko, Zero"),
    ("03", "RESTROOM", "3070 16-Ga HM Door, 5-7/8\" HMD Frame, Ives 5BB1 Hinges (3ea), LCN 4040XP, Schlage B57 Indicator Deadbolt, Ives Push/Pull/Kick/Stop", 1, "EA", 961.66, 961.66, "catalog_list_x_multiplier", "[schedule] Sheet A2.2 / Ives, LCN, Schlage"),
    ("04", "TOWER ACCESS", "3030 JL Industries XPA-3636H4W-R 36\"x36\" Roof Access Panel w/ Flush Handle & Rainhood", 1, "EA", 383.56, 383.56, "manual_wholesaler_net", "[schedule] Sheet A2.2 / JL Industries")
]

door_start_row = row_idx
for tag, room, desc, qty, unit, unit_sale, ext_sale, basis, citation in door_rows:
    ws.cell(row=row_idx, column=1, value=tag).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=2, value=room).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=3, value=desc).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=4, value=qty).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=5, value=unit).alignment = Alignment(horizontal="center", vertical="center")
    
    c6 = ws.cell(row=row_idx, column=6, value=unit_sale)
    c6.number_format = "$#,##0.00"
    c6.alignment = Alignment(horizontal="right", vertical="center")
    
    c7 = ws.cell(row=row_idx, column=7, value=f"=D{row_idx}*F{row_idx}")
    c7.number_format = "$#,##0.00"
    c7.alignment = Alignment(horizontal="right", vertical="center")
    
    ws.cell(row=row_idx, column=8, value=basis).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=9, value=citation).alignment = Alignment(horizontal="left", vertical="center")
    
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).font = regular_font
        ws.cell(row=row_idx, column=c).border = thin_border
    ws.row_dimensions[row_idx].height = 22
    row_idx += 1

door_end_row = row_idx - 1

# Door Subtotal Row
ws.cell(row=row_idx, column=3, value="Subtotal Division 08 Doors & Hardware:").font = bold_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
sub_cell = ws.cell(row=row_idx, column=7, value=f"=SUM(G{door_start_row}:G{door_end_row})")
sub_cell.font = bold_font
sub_cell.number_format = "$#,##0.00"
sub_cell.alignment = Alignment(horizontal="right", vertical="center")
for c in range(1, 10):
    ws.cell(row=row_idx, column=c).fill = subtotal_fill
    ws.cell(row=row_idx, column=c).border = subtotal_border
ws.row_dimensions[row_idx].height = 22
row_idx += 2

# Section 2: Division 10 Washroom Accessories
add_section_header("DIVISION 10 — WASHROOM ACCESSORIES")

acc_rows = [
    ("PA51", "RESTROOM", "Bobrick B-5806 1-1/2\" Dia Stainless Steel ADA Grab Bar (36\" / 42\")", 2, "EA", 72.95, 145.90, "catalog_list_x_multiplier", "[schedule] Sheet A1.1 / Bobrick p.16"),
    ("PA52", "RESTROOM", "Bobrick B-165-1836 Stainless Steel Framed Mirror 18\" x 36\"", 1, "EA", 104.66, 104.66, "catalog_list_x_multiplier", "[schedule] Sheet A1.1 / Bobrick p.5"),
    ("PA64", "RESTROOM", "Bobrick B-254 Surface Mounted Sanitary Napkin Disposal Unit", 1, "EA", 110.23, 110.23, "catalog_list_x_multiplier", "[schedule] Sheet A1.1 / Bobrick"),
    ("PA65", "RESTROOM", "Bobrick B-221 Surface Mounted Toilet Seat Cover Dispenser", 1, "EA", 56.36, 56.36, "catalog_list_x_multiplier", "[schedule] Sheet A1.1 / Bobrick"),
    ("PA66", "RESTROOM", "Bobrick B-212 Clothes Hook with Bumper", 1, "EA", 20.91, 20.91, "catalog_list_x_multiplier", "[schedule] Sheet A1.1 / Bobrick")
]

acc_start_row = row_idx
for tag, room, desc, qty, unit, unit_sale, ext_sale, basis, citation in acc_rows:
    ws.cell(row=row_idx, column=1, value=tag).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=2, value=room).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=3, value=desc).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=4, value=qty).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=5, value=unit).alignment = Alignment(horizontal="center", vertical="center")
    
    c6 = ws.cell(row=row_idx, column=6, value=unit_sale)
    c6.number_format = "$#,##0.00"
    c6.alignment = Alignment(horizontal="right", vertical="center")
    
    c7 = ws.cell(row=row_idx, column=7, value=f"=D{row_idx}*F{row_idx}")
    c7.number_format = "$#,##0.00"
    c7.alignment = Alignment(horizontal="right", vertical="center")
    
    ws.cell(row=row_idx, column=8, value=basis).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=9, value=citation).alignment = Alignment(horizontal="left", vertical="center")
    
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).font = regular_font
        ws.cell(row=row_idx, column=c).border = thin_border
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

acc_end_row = row_idx - 1

# Accessories Subtotal Row
ws.cell(row=row_idx, column=3, value="Subtotal Division 10 Washroom Accessories:").font = bold_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
sub_cell = ws.cell(row=row_idx, column=7, value=f"=SUM(G{acc_start_row}:G{acc_end_row})")
sub_cell.font = bold_font
sub_cell.number_format = "$#,##0.00"
sub_cell.alignment = Alignment(horizontal="right", vertical="center")
for c in range(1, 10):
    ws.cell(row=row_idx, column=c).fill = subtotal_fill
    ws.cell(row=row_idx, column=c).border = subtotal_border
ws.row_dimensions[row_idx].height = 22
row_idx += 2

# Section 3: Division 06 FRP Wall Panels & Trims
add_section_header("DIVISION 06 64 — FRP WALL PANELS & MOULDINGS (PROVISIONAL)")

frp_rows = [
    ("FRP-1", "KITCHEN/PREP", ".090\" Class C FRP Wall Panel 4'x10' White Embossed", 40, "SHT", 57.53, 2301.20, "catalog_list_x_multiplier", "[spec] Sheet A2.1 / NUDO"),
    ("TRIM-1", "KITCHEN/PREP", "10ft PVC Division Bar Moulding", 39, "STK", 8.90, 347.10, "catalog_list_x_multiplier", "[spec] Sheet A2.1 / NUDO"),
    ("TRIM-2", "KITCHEN/PREP", "10ft PVC Inside Corner Moulding", 10, "STK", 8.90, 89.00, "catalog_list_x_multiplier", "[spec] Sheet A2.1 / NUDO"),
    ("TRIM-3", "KITCHEN/PREP", "10ft PVC Outside Corner Moulding", 4, "STK", 8.90, 35.60, "catalog_list_x_multiplier", "[spec] Sheet A2.1 / NUDO"),
    ("TRIM-4", "KITCHEN/PREP", "10ft PVC End Cap J-Moulding", 30, "STK", 7.53, 225.90, "catalog_list_x_multiplier", "[spec] Sheet A2.1 / NUDO"),
    ("ADH-1", "KITCHEN/PREP", "FRP Water-Based Adhesive 4-Gallon Pails", 6, "PAIL", 93.15, 558.90, "catalog_list_x_multiplier", "[spec] Sheet A2.1 / NUDO")
]

frp_start_row = row_idx
for tag, room, desc, qty, unit, unit_sale, ext_sale, basis, citation in frp_rows:
    ws.cell(row=row_idx, column=1, value=tag).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=2, value=room).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=3, value=desc).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=4, value=qty).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=5, value=unit).alignment = Alignment(horizontal="center", vertical="center")
    
    c6 = ws.cell(row=row_idx, column=6, value=unit_sale)
    c6.number_format = "$#,##0.00"
    c6.alignment = Alignment(horizontal="right", vertical="center")
    
    c7 = ws.cell(row=row_idx, column=7, value=f"=D{row_idx}*F{row_idx}")
    c7.number_format = "$#,##0.00"
    c7.alignment = Alignment(horizontal="right", vertical="center")
    
    ws.cell(row=row_idx, column=8, value=basis).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_idx, column=9, value=citation).alignment = Alignment(horizontal="left", vertical="center")
    
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).font = regular_font
        ws.cell(row=row_idx, column=c).border = thin_border
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

frp_end_row = row_idx - 1

# FRP Subtotal Row
ws.cell(row=row_idx, column=3, value="Subtotal Division 06 FRP Panels & Trims:").font = bold_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
sub_cell = ws.cell(row=row_idx, column=7, value=f"=SUM(G{frp_start_row}:G{frp_end_row})")
sub_cell.font = bold_font
sub_cell.number_format = "$#,##0.00"
sub_cell.alignment = Alignment(horizontal="right", vertical="center")
for c in range(1, 10):
    ws.cell(row=row_idx, column=c).fill = subtotal_fill
    ws.cell(row=row_idx, column=c).border = subtotal_border
ws.row_dimensions[row_idx].height = 22
row_idx += 2

# Summary Block Header
add_section_header("QUOTATION SUMMARY & TERMS")

# Write Summary Table
sum_start_row = row_idx
ws.cell(row=row_idx, column=3, value="Base Bid Material Subtotal:").font = bold_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
c = ws.cell(row=row_idx, column=7, value=f"=G{door_end_row+1}+G{acc_end_row+1}+G{frp_end_row+1}")
c.font = bold_font
c.number_format = "$#,##0.00"
c.alignment = Alignment(horizontal="right", vertical="center")
for col in range(1, 10):
    ws.cell(row=row_idx, column=col).border = thin_border
    ws.cell(row=row_idx, column=col).fill = subtotal_fill
ws.row_dimensions[row_idx].height = 22
row_idx += 1

ws.cell(row=row_idx, column=3, value="Freight Charge:").font = regular_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
c = ws.cell(row=row_idx, column=7, value="TBD - Excluded at Estimate Stage")
c.font = italic_font
c.alignment = Alignment(horizontal="right", vertical="center")
for col in range(1, 10):
    ws.cell(row=row_idx, column=col).border = thin_border
ws.row_dimensions[row_idx].height = 20
row_idx += 1

ws.cell(row=row_idx, column=3, value="Sales Tax (PA 0.0% Supply-Only):").font = regular_font
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
c = ws.cell(row=row_idx, column=7, value=0.00)
c.font = regular_font
c.number_format = "$#,##0.00"
c.alignment = Alignment(horizontal="right", vertical="center")
for col in range(1, 10):
    ws.cell(row=row_idx, column=col).border = thin_border
ws.row_dimensions[row_idx].height = 20
row_idx += 1

# Grand Total Row
ws.cell(row=row_idx, column=3, value="GRAND TOTAL MATERIAL QUOTATION:").font = Font(name="Calibri", size=12, bold=True, color="1A365D")
ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right", vertical="center")
c = ws.cell(row=row_idx, column=7, value=f"=G{sum_start_row}+G{sum_start_row+2}")
c.font = Font(name="Calibri", size=12, bold=True, color="1A365D")
c.number_format = "$#,##0.00"
c.alignment = Alignment(horizontal="right", vertical="center")
for col in range(1, 10):
    ws.cell(row=row_idx, column=col).fill = total_fill
    ws.cell(row=row_idx, column=col).border = total_border
ws.row_dimensions[row_idx].height = 26
row_idx += 2

# Terms and Footers
ws.cell(row=row_idx, column=1, value="STANDARD COMMERCIAL TERMS & ESTIMATING NOTES:").font = bold_font
row_idx += 1
terms = [
    "1. Quotation is supply-only material F.O.B. factory / CBC warehouse. Installation by others.",
    "2. Material quantities derived strictly from drawing schedules on 1_Architectural.pdf.",
    "3. Hardware pricing reflects specified manufacturers (Ives, LCN, Alarm Lock, Von Duprin, Schlage, Pemko, Zero).",
    "4. Restroom accessories priced directly against Bobrick 2020 Net Price List.",
    "5. Quote valid for 30 days from date of issue. Subject to Hamilton Parker credit approval."
]
for t in terms:
    ws.cell(row=row_idx, column=1, value=t).font = italic_font
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1

# Auto-fit Column Widths
col_widths = {
    "A": 14,
    "B": 22,
    "C": 75,
    "D": 8,
    "E": 8,
    "F": 16,
    "G": 16,
    "H": 28,
    "I": 50
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

wb.save(output_path)
print(f"Quotation excel workbook saved successfully to {output_path}")
