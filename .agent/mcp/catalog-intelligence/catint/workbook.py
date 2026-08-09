"""Parse a vendor price list held as a spreadsheet into the same product rows a PDF
price book yields.

Roughly half this shelf's pricing never reaches the index because it arrives as .xlsx,
not .pdf - World Dryer's whole book, Gamco's distributor list, and Hamilton Parker's own
net-cost sheets. `cbc_engine.engine` even records World Dryer as "Price book is Excel
(not indexed by catalog-intelligence)".

These are far easier to read than the PDFs: a header row and typed cells, with none of
the column-x-position inference `index.parse_products` needs. Two things still have to
be got right:

- **A sheet can print several prices per model.** Hamilton Parker's Bobrick sheet has
  LIST, DEALER, COMMERCIAL, MAP and HAMILTON PARKER NET side by side. Each becomes its
  own row carrying its own `price_basis`, because collapsing them to one number loses
  the only thing that says whether a multiplier still has to be applied.
- **Headers are written by hand.** The Gamco sheets say "Distibutor NET", misspelled.
  Matching is on normalized substrings, never equality.
"""

import os
import re

# Header text that names the product itself, most specific first - World Dryer prints
# both "Model Family" and "Model #", and only the second is orderable.
MODEL_HEADERS = ("model number", "model #", "model no", "part number", "part #",
                 "part no", "item number", "item #", "item no", "catalog number",
                 "catalog #", "sku", "model", "part", "item")

# ...and header text that looks like a model column but names a grouping. Taking one of
# these as the model gives every row in the group the same non-orderable code.
MODEL_EXCLUDE = ("family", "group", "series", "type", "category", "line")

# Header text that names a price column, mapped to the basis recorded on the row.
# Order matters: the first substring found wins, so the specific ones lead.
PRICE_HEADERS = (
    ("hamilton parker net", "hamilton_parker_net"),
    ("distibutor net", "distributor_net"),      # sic - as printed on the Gamco sheets
    ("distributor net", "distributor_net"),
    ("dealer net", "dealer_net"),
    ("net cost", "net"),
    ("net price", "net"),
    ("list price", "list"),
    ("map", "map"),
    ("dealer", "dealer"),
    ("commercial", "commercial"),
    ("list", "list"),
    ("net", "net"),
    ("price", "list"),
    ("cost", "net"),
)

# A basis that is already a cost, so a vendor multiplier must NOT be applied on top.
NET_BASES = {"net", "dealer_net", "distributor_net", "hamilton_parker_net"}

_MONEY_RX = re.compile(r"^\$?\s*-?[\d,]+(?:\.\d+)?$")
_WS = re.compile(r"\s+")


def _norm(value):
    """A cell as comparable lowercase text."""
    return _WS.sub(" ", str(value if value is not None else "")).strip().lower()


def _money(value):
    """A cell as a price, or None. Excel gives floats; humans give '$1,058.00'."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 4) if value > 0 else None
    text = str(value or "").strip()
    if not text or not _MONEY_RX.match(text):
        return None
    try:
        number = float(text.replace("$", "").replace(",", "").strip())
    except ValueError:
        return None
    return round(number, 4) if number > 0 else None


def _price_basis(header):
    for needle, basis in PRICE_HEADERS:
        if needle in header:
            return basis
    return None


def classify_headers(cells):
    """(model_col, {price_col: basis}, [descriptive_cols]) for a candidate header row.

    A row is only a header if it names both a product column and at least one price
    column - otherwise a title banner or a section break would be read as one.
    """
    headers = [(i, _norm(cell)) for i, cell in enumerate(cells) if _norm(cell)]

    # Most specific model header wins, wherever it sits: "Model #" in column 5 beats
    # "Model Family" in column 0.
    model_col = None
    for needle in MODEL_HEADERS:
        for i, header in headers:
            if needle in header and not any(x in header for x in MODEL_EXCLUDE):
                model_col = i
                break
        if model_col is not None:
            break

    prices, descriptive = {}, []
    for i, header in headers:
        if i == model_col:
            continue
        basis = _price_basis(header)
        if basis:
            prices[i] = basis
        else:
            descriptive.append((i, header))
    return model_col, prices, descriptive


def find_header(rows, limit=25):
    """Index of the header row, or None. Sheets often carry a title above it."""
    for i, cells in enumerate(rows[:limit]):
        model_col, prices, _ = classify_headers(cells)
        if model_col is not None and prices:
            return i
    return None


def _read_sheet(worksheet, max_rows=20000):
    out = []
    for row in worksheet.iter_rows(values_only=True):
        out.append(list(row))
        if len(out) >= max_rows:
            break
    return out


def parse_sheet(rows, sheet_name=""):
    """Product rows from one already-read worksheet."""
    header_at = find_header(rows)
    if header_at is None:
        return []
    model_col, price_cols, descriptive = classify_headers(rows[header_at])
    labels = {i: _norm(rows[header_at][i]) for i in price_cols}

    out = []
    section = sheet_name
    for cells in rows[header_at + 1:]:
        if model_col >= len(cells):
            continue
        model = str(cells[model_col] or "").strip()
        prices = {i: _money(cells[i]) for i in price_cols if i < len(cells)}

        # A row with a label but no price is a section break ("SOAP DISPENSERS"),
        # which is worth keeping as the section for the rows beneath it.
        if model and not any(v is not None for v in prices.values()):
            if len(model) > 2 and not any(ch.isdigit() for ch in model):
                section = model
            continue
        if not model:
            continue

        parts = [str(cells[i]).strip() for i, _ in descriptive
                 if i < len(cells) and cells[i] not in (None, "")]
        description = " - ".join(dict.fromkeys(parts)) or model

        for col, price in prices.items():
            if price is None:
                continue
            out.append({
                "model": model,
                "description": description,
                "size": "",
                "finish": "",
                "list_price": price,
                "qty": "",
                "disc_code": "",
                "price_basis": price_cols[col],
                "price_column": labels[col],
                "section": section,
            })
    return out


def parse_workbook(path):
    """[(sheet_name, [product rows])] for every worksheet that holds a price table."""
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out = []
        for worksheet in book.worksheets:
            rows = parse_sheet(_read_sheet(worksheet), worksheet.title)
            if rows:
                out.append((worksheet.title, rows))
        return out
    finally:
        book.close()


# --------------------------------------------------------------------------- #
# Crossover sheets
# --------------------------------------------------------------------------- #

# Header words that mark a column of *another vendor's* equivalent model rather than a
# price. A sheet built from these is a crossover table, not a price list.
CROSSOVER_HEADERS = ("equivalent", "crossover", "cross over", "cross reference",
                     "cross-reference", "equal", "comparable")

_NOT_A_VALUE = {"", "none", "n/a", "na", "-", "--", "n\\a"}


def _vendor_from_header(header):
    """'Bobrick Equivalent' -> 'BOBRICK'. 'Bradley Crossover' -> 'BRADLEY'."""
    name = header
    for word in CROSSOVER_HEADERS:
        name = name.replace(word, " ")
    return " ".join(name.split()).upper()


def _cell_value(cell):
    text = str(cell).strip() if cell is not None else ""
    return "" if text.lower() in _NOT_A_VALUE else text


def parse_crossover(rows, source_vendor, limit=25):
    """Equivalence classes from a crossover sheet: [[(vendor, model), ...], ...].

    Returned as classes, not as pairs. The Shandas sheet expands to 3,162 directed
    pairs, and storing those would be storing the same 409 facts a dozen times over.
    """
    header_at, columns = None, []
    for i, cells in enumerate(rows[:limit]):
        found = []
        for j, cell in enumerate(cells):
            header = _norm(cell)
            if not header:
                continue
            if any(w in header for w in CROSSOVER_HEADERS):
                found.append((j, _vendor_from_header(header) or "UNKNOWN"))
            elif not found and any(h in header for h in MODEL_HEADERS):
                found.append((j, source_vendor.upper()))
        # Two or more vendors, and no price column - otherwise it is a price list.
        if len(found) >= 2 and not any(
                _price_basis(_norm(c)) for c in cells if _norm(c)):
            header_at, columns = i, found
            break
    if header_at is None:
        return []

    classes = []
    for cells in rows[header_at + 1:]:
        members = []
        for j, vendor in columns:
            if j < len(cells):
                model = _cell_value(cells[j])
                if model:
                    members.append((vendor, model))
        if len(members) >= 2:          # a row naming one vendor states no equivalence
            classes.append(members)
    return classes


def parse_workbook_crossovers(path, source_vendor):
    """[(sheet_name, [equivalence class, ...])] for every crossover sheet."""
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out = []
        for worksheet in book.worksheets:
            classes = parse_crossover(_read_sheet(worksheet), source_vendor)
            if classes:
                out.append((worksheet.title, classes))
        return out
    finally:
        book.close()


SUPPORTED = (".xlsx", ".xlsm")


def is_workbook(path):
    return os.path.splitext(path)[1].lower() in SUPPORTED
